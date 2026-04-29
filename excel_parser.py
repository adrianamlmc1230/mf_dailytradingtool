"""Excel anchor parser utilities for the local Streamlit tool."""

from __future__ import annotations

from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

ANCHOR_TYPES = {"前", "尾"}
BLOCK_HEIGHT = 10


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_anchor_workbook(workbook_source: str | bytes | BinaryIO, source_name: str) -> dict[str, Any]:
    workbook = load_workbook(workbook_source, data_only=True)

    blocks: list[dict[str, Any]] = []
    team_pool_rows: list[dict[str, str]] = []
    unique_values: dict[str, None] = {}
    stats = {
        "front_anchor_count": 0,
        "tail_anchor_count": 0,
        "block_count": 0,
        "non_empty_text_count": 0,
        "unique_text_count": 0,
    }
    warnings_list: list[str] = []

    source_label = str(source_name or "uploaded.xlsx")

    for worksheet in workbook.worksheets:
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                anchor_type = normalize_text(cell.value)
                if anchor_type not in ANCHOR_TYPES:
                    continue

                if anchor_type == "前":
                    stats["front_anchor_count"] += 1
                else:
                    stats["tail_anchor_count"] += 1

                extracted_values: list[str] = []
                missing_rows = []
                for offset in range(1, BLOCK_HEIGHT + 1):
                    target_row = row + offset
                    if target_row > worksheet.max_row:
                        missing_rows.append(target_row)
                        continue

                    target_value = normalize_text(worksheet.cell(row=target_row, column=col).value)
                    if target_value:
                        extracted_values.append(target_value)
                        unique_values.setdefault(target_value, None)
                        stats["non_empty_text_count"] += 1
                        team_pool_rows.append(
                            {
                                "source_file": source_label,
                                "sheet_name": worksheet.title,
                                "anchor_type": anchor_type,
                                "anchor_cell": cell.coordinate,
                                "team_name": target_value,
                            }
                        )

                if missing_rows:
                    warnings_list.append(
                        f"Warning: {worksheet.title} {cell.coordinate} below-block has fewer than 10 cells "
                        f"(missing rows starting at {missing_rows[0]})."
                    )

                if not extracted_values:
                    warnings_list.append(
                        f"Warning: {worksheet.title} {cell.coordinate} below-block 10 cells are all empty."
                    )

                blocks.append(
                    {
                        "sheet_name": worksheet.title,
                        "anchor_type": anchor_type,
                        "anchor_cell": cell.coordinate,
                        "extracted_values": extracted_values,
                    }
                )
                stats["block_count"] += 1

    stats["unique_text_count"] = len(unique_values)
    return {
        "source_file": source_label,
        "blocks": blocks,
        "team_pool_rows": team_pool_rows,
        "unique_values": list(unique_values.keys()),
        "stats": stats,
        "warnings": warnings_list,
    }
